from __future__ import annotations

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO
import re
from decimal import Decimal
from typing import Optional, List

from pydantic import BaseModel, Field

from app.core.db import get_db
from app.models.requirements import Requirement

router = APIRouter(prefix="/requirements", tags=["requirements"])

WEEKDAY_MAP = {
    "lunes": 0,
    "martes": 1,
    "miercoles": 2, "miércoles": 2,
    "jueves": 3,
    "viernes": 4,
    "sabado": 5, "sábado": 5,
    "domingo": 6,
}

H_COL_RE = re.compile(r"^H\d{4}$")


def hcol_to_min(col: str) -> int:
    hh = int(col[1:3])
    mm = int(col[3:5])
    return hh * 60 + mm


# -----------------------
# GET /requirements (lista)
# -----------------------
class RequirementOut(BaseModel):
    campaign_id: int
    period: int
    weekday: int = Field(ge=0, le=6)
    minute: int = Field(ge=0, le=1430)
    required: Decimal

    class Config:
        from_attributes = True  # pydantic v2


@router.get("", response_model=list[RequirementOut])
def list_requirements(
    campaign_id: int = Query(..., ge=1),
    period: int = Query(...),
    weekday: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
):
    q = db.query(Requirement).filter(
        Requirement.campaign_id == campaign_id,
        Requirement.period == period,
    )
    if weekday is not None:
        q = q.filter(Requirement.weekday == weekday)

    rows = q.order_by(Requirement.weekday.asc(), Requirement.minute.asc()).all()
    return rows


# -----------------------
# GET /requirements/series (48 slots)
# -----------------------
class RequirementSeriesOut(BaseModel):
    campaign_id: int
    period: int
    weekday: int
    minutes: List[int]
    required: List[Optional[Decimal]]

class RequirementWeekSeriesOut(BaseModel):
    campaign_id: int
    period: int
    minutes: List[int]
    weekdays: List[int]                 # [0..6]
    series: List[RequirementSeriesOut]  # una por weekday


@router.get("/series", response_model=RequirementSeriesOut)
def requirements_series(
    campaign_id: int = Query(..., ge=1),
    period: int = Query(...),
    weekday: int = Query(..., ge=0, le=6),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Requirement)
        .filter(
            Requirement.campaign_id == campaign_id,
            Requirement.period == period,
            Requirement.weekday == weekday,
        )
        .order_by(Requirement.minute.asc())
        .all()
    )

    minutes = list(range(0, 24 * 60, 30))  # 0..1410
    by_minute = {r.minute: r.required for r in rows}
    series = [by_minute.get(m) for m in minutes]

    return RequirementSeriesOut(
        campaign_id=campaign_id,
        period=period,
        weekday=weekday,
        minutes=minutes,
        required=series,
    )

@router.get("/series/week", response_model=RequirementWeekSeriesOut)
def requirements_series_week(
    campaign_id: int = Query(..., ge=1),
    period: int = Query(...),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Requirement)
        .filter(
            Requirement.campaign_id == campaign_id,
            Requirement.period == period,
        )
        .order_by(Requirement.weekday.asc(), Requirement.minute.asc())
        .all()
    )

    minutes = list(range(0, 24 * 60, 30))  # 0..1410
    by_wd_min = {(r.weekday, r.minute): r.required for r in rows}

    series_list: List[RequirementSeriesOut] = []
    for wd in range(7):
        reqs = [by_wd_min.get((wd, m)) for m in minutes]
        series_list.append(
            RequirementSeriesOut(
                campaign_id=campaign_id,
                period=period,
                weekday=wd,
                minutes=minutes,
                required=reqs,
            )
        )

    return RequirementWeekSeriesOut(
        campaign_id=campaign_id,
        period=period,
        minutes=minutes,
        weekdays=list(range(7)),
        series=series_list,
    )

# -----------------------
# POST /requirements/import
# -----------------------
@router.post("/import")
async def import_requirements(
    campaign_id: int = Form(...),
    period: int = Form(...),
    sheet_name: str | None = Form(None),
    sheet_index: int = Form(0),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"sheet_name '{sheet_name}' not found")
        ws = wb[sheet_name]
    else:
        if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
            raise HTTPException(status_code=400, detail=f"sheet_index out of range (0..{len(wb.sheetnames)-1})")
        ws = wb[wb.sheetnames[sheet_index]]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if not header or "Proceso" not in header or "Día" not in header:
        raise HTTPException(status_code=400, detail="Header inválido: debe contener 'Proceso' y 'Día'")

    col_idx = {name: i for i, name in enumerate(header) if isinstance(name, str)}
    proceso_i = col_idx.get("Proceso")
    dia_i = col_idx.get("Día")

    h_cols = [(name, i) for name, i in col_idx.items() if H_COL_RE.match(name)]
    if not h_cols:
        raise HTTPException(status_code=400, detail="No se encontraron columnas H0000..H2330")

    h_cols.sort(key=lambda x: hcol_to_min(x[0]))

    inserted = 0
    updated = 0
    weekday_rows: dict[int, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        proceso = (row[proceso_i] or "")
        dia = (row[dia_i] or "")

        if not isinstance(proceso, str) or not isinstance(dia, str):
            continue

        if proceso.strip().lower() != "generales":
            continue

        dia_norm = dia.strip().lower()
        if dia_norm not in WEEKDAY_MAP:
            continue

        weekday = WEEKDAY_MAP[dia_norm]
        count_slots = 0

        for hname, hi in h_cols:
            val = row[hi]
            if val is None or val == "":
                continue

            minute = hcol_to_min(hname)

            try:
                req = Decimal(str(val))
            except Exception:
                raise HTTPException(status_code=400, detail=f"Valor inválido en {hname} ({dia}): {val}")

            existing = (
                db.query(Requirement)
                .filter(
                    Requirement.campaign_id == campaign_id,
                    Requirement.period == period,
                    Requirement.weekday == weekday,
                    Requirement.minute == minute,
                )
                .one_or_none()
            )

            if existing is None:
                db.add(
                    Requirement(
                        campaign_id=campaign_id,
                        period=period,
                        weekday=weekday,
                        minute=minute,
                        required=req,
                    )
                )
                inserted += 1
            else:
                if existing.required != req:
                    existing.required = req
                    updated += 1

            count_slots += 1

        weekday_rows[weekday] = weekday_rows.get(weekday, 0) + count_slots

    db.commit()

    return {
        "sheet": ws.title,
        "campaign_id": campaign_id,
        "period": period,
        "inserted": inserted,
        "updated": updated,
        "weekday_rows": weekday_rows,
    }