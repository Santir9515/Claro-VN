# backend/app/main.py
from __future__ import annotations

from typing import Optional

from datetime import date as date_type
from decimal import Decimal
from io import BytesIO
import re

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy import and_
from sqlalchemy.orm import Session, aliased
from openpyxl import load_workbook

from app.core.db import get_db
from app.core.timeutils import hhmm_to_min, min_to_hhmm

from app.models.campaign import Campaign
from app.models.advisor import Advisor
from app.models.shift import Shift
from app.models.absence import Absence
from app.models.requirements import Requirement

from app.api.routes.requirements import router as requirements_router


from app.api.schemas import (
    ShiftUpsertIn,
    ShiftOut,
    AbsenceUpsertIn,
    AbsenceOut,
    RosterRow,
)

from app.workers.break_tasks import assign_break

app = FastAPI(title="WFM Breaks MVP")
app.include_router(requirements_router)


@app.get("/health")
def health():
    return {"status": "ok"}


# -----------------------
# CAMPAIGNS
# -----------------------
class CampaignIn(BaseModel):
    name: str


@app.post("/campaigns")
def create_campaign(payload: CampaignIn, db: Session = Depends(get_db)):
    c = Campaign(name=payload.name)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name}


@app.get("/campaigns")
def list_campaigns(db: Session = Depends(get_db)):
    rows = db.query(Campaign).order_by(Campaign.id.asc()).all()
    return [{"id": r.id, "name": r.name} for r in rows]


# -----------------------
# ADVISORS
# -----------------------
class AdvisorIn(BaseModel):
    name: str
    campaign_id: int


@app.post("/advisors")
def create_advisor(payload: AdvisorIn, db: Session = Depends(get_db)):
    camp = db.query(Campaign).filter(Campaign.id == payload.campaign_id).one_or_none()
    if camp is None:
        raise HTTPException(status_code=400, detail="campaign_id not found")

    a = Advisor(name=payload.name, campaign_id=payload.campaign_id)
    db.add(a)
    db.commit()
    db.refresh(a)
    return {"id": a.id, "name": a.name, "campaign_id": a.campaign_id}


@app.get("/advisors")
def list_advisors(campaign_id: int | None = None, db: Session = Depends(get_db)):
    q = db.query(Advisor)
    if campaign_id is not None:
        q = q.filter(Advisor.campaign_id == campaign_id)
    rows = q.order_by(Advisor.id.asc()).all()
    return [{"id": r.id, "name": r.name, "campaign_id": r.campaign_id} for r in rows]


# -----------------------
# SHIFTS (UPSERT + LIST)
# -----------------------
@app.post("/shifts", response_model=ShiftOut)
def upsert_shift(payload: ShiftUpsertIn, db: Session = Depends(get_db)):
    # valida y convierte HH:MM -> minutos
    try:
        start_min = hhmm_to_min(payload.start)
        end_min = hhmm_to_min(payload.end)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if end_min <= start_min:
        raise HTTPException(status_code=400, detail="end must be greater than start")

    # valida FK advisor
    adv = db.query(Advisor).filter(Advisor.id == payload.advisor_id).one_or_none()
    if adv is None:
        raise HTTPException(status_code=400, detail="advisor_id not found")

    row = (
        db.query(Shift)
        .filter(Shift.advisor_id == payload.advisor_id, Shift.day == payload.day)
        .one_or_none()
    )

    if row is None:
        row = Shift(
            advisor_id=payload.advisor_id,
            day=payload.day,
            start_minute=start_min,
            end_minute=end_min,
        )
        db.add(row)
    else:
        row.start_minute = start_min
        row.end_minute = end_min

    db.commit()

    # Encolar asignación de break +30 min (1800s)
    assign_break.apply_async(
        args=[payload.advisor_id, payload.day.isoformat()],
        countdown=1800,
    )

    return ShiftOut(
        advisor_id=payload.advisor_id,
        day=payload.day,
        start=min_to_hhmm(start_min),
        end=min_to_hhmm(end_min),
    )


@app.get("/shifts", response_model=list[ShiftOut])
def list_shifts(
    day: Optional[date_type] = None,
    advisor_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Shift)
    if day is not None:
        q = q.filter(Shift.day == day)
    if advisor_id is not None:
        q = q.filter(Shift.advisor_id == advisor_id)

    rows = q.order_by(Shift.day.asc(), Shift.advisor_id.asc()).all()
    return [
        ShiftOut(
            advisor_id=r.advisor_id,
            day=r.day,
            start=min_to_hhmm(r.start_minute),
            end=min_to_hhmm(r.end_minute),
        )
        for r in rows
    ]


# -----------------------
# ABSENCES (UPSERT + LIST)
# -----------------------
@app.post("/absences", response_model=AbsenceOut)
def upsert_absence(payload: AbsenceUpsertIn, db: Session = Depends(get_db)):
    # valida FK advisor
    adv = db.query(Advisor).filter(Advisor.id == payload.advisor_id).one_or_none()
    if adv is None:
        raise HTTPException(status_code=400, detail="advisor_id not found")

    row = (
        db.query(Absence)
        .filter(Absence.advisor_id == payload.advisor_id, Absence.day == payload.day)
        .one_or_none()
    )

    if row is None:
        row = Absence(
            advisor_id=payload.advisor_id,
            day=payload.day,
            is_absent=payload.is_absent,
        )
        db.add(row)
    else:
        row.is_absent = payload.is_absent

    db.commit()

    return AbsenceOut(
        advisor_id=payload.advisor_id,
        day=payload.day,
        is_absent=row.is_absent,
    )


@app.get("/absences", response_model=list[AbsenceOut])
def list_absences(
    day: Optional[date_type] = None,
    advisor_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Absence)
    if day is not None:
        q = q.filter(Absence.day == day)
    if advisor_id is not None:
        q = q.filter(Absence.advisor_id == advisor_id)

    rows = q.order_by(Absence.day.asc(), Absence.advisor_id.asc()).all()
    return [AbsenceOut(advisor_id=r.advisor_id, day=r.day, is_absent=r.is_absent) for r in rows]


# -----------------------
# ROSTER (LEFT JOIN)
# -----------------------
@app.get("/roster", response_model=list[RosterRow])
def roster(day: date_type, db: Session = Depends(get_db)):
    Sh = aliased(Shift)
    Ab = aliased(Absence)

    rows = (
        db.query(Advisor, Sh, Ab)
        .outerjoin(Sh, and_(Sh.advisor_id == Advisor.id, Sh.day == day))
        .outerjoin(Ab, and_(Ab.advisor_id == Advisor.id, Ab.day == day))
        .order_by(Advisor.id.asc())
        .all()
    )

    out: list[RosterRow] = []
    for adv, sh, ab in rows:
        out.append(
            RosterRow(
                advisor_id=adv.id,
                advisor_name=adv.name,
                day=day,
                shift_start=min_to_hhmm(sh.start_minute) if sh else None,
                shift_end=min_to_hhmm(sh.end_minute) if sh else None,
                is_absent=bool(ab.is_absent) if ab else False,
            )
        )
    return out


# -----------------------
# REQUIREMENTS (IMPORT + LIST)
# -----------------------
WEEKDAY_MAP = {
    "lunes": 0,
    "martes": 1,
    "miercoles": 2,
    "miércoles": 2,
    "jueves": 3,
    "viernes": 4,
    "sabado": 5,
    "sábado": 5,
    "domingo": 6,
}

H_COL_RE = re.compile(r"^H\d{4}$")


def hcol_to_min(col: str) -> int:
    hh = int(col[1:3])
    mm = int(col[3:5])
    return hh * 60 + mm


@app.post("/requirements/import")
async def import_requirements(
    campaign_id: int = Form(...),
    period: int = Form(...),
    sheet_name: str | None = Form(None),
    sheet_index: int = Form(0),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # valida campaña
    camp = db.query(Campaign).filter(Campaign.id == campaign_id).one_or_none()
    if camp is None:
        raise HTTPException(status_code=400, detail="campaign_id not found")

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"sheet_name '{sheet_name}' not found")
        ws = wb[sheet_name]
    else:
        if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
            raise HTTPException(status_code=400, detail=f"sheet_index out of range (0..{len(wb.sheetnames)-1})")
        ws = wb[wb.sheetnames[sheet_index]]

    # Buscar header real (muchos Excels tienen filas vacías arriba)
    header_row_idx = None
    header = None
    for r in range(1, 30):
        row_vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
        if row_vals and "Proceso" in row_vals and ("Día" in row_vals or "Dia" in row_vals):
            header_row_idx = r
            header = row_vals
            break

    if header_row_idx is None or header is None:
        raise HTTPException(status_code=400, detail="No encontré header con 'Proceso' y 'Día' en las primeras filas")

    # Normalizar posibles variantes de "Día"
    header = ["Día" if v == "Dia" else v for v in header]

    col_idx = {name: i for i, name in enumerate(header) if isinstance(name, str)}
    proceso_i = col_idx.get("Proceso")
    dia_i = col_idx.get("Día")

    if proceso_i is None or dia_i is None:
        raise HTTPException(status_code=400, detail="Header inválido: debe contener 'Proceso' y 'Día'")

    h_cols = [(name, i) for name, i in col_idx.items() if H_COL_RE.match(name)]
    if not h_cols:
        raise HTTPException(status_code=400, detail="No se encontraron columnas H0000..H2330")

    h_cols.sort(key=lambda x: hcol_to_min(x[0]))

    inserted = 0
    updated = 0
    weekday_slots: dict[int, int] = {}

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        proceso = (row[proceso_i] or "")
        dia = (row[dia_i] or "")

        if not isinstance(proceso, str) or not isinstance(dia, str):
            continue

        if proceso.strip().lower() != "generales":
            continue

        dia_norm = dia.strip().lower()
        if dia_norm not in WEEKDAY_MAP:
            continue

        weekday = WEEKDAY_MAP[dia_norm]
        count_slots = 0

        for hname, hi in h_cols:
            val = row[hi] if hi < len(row) else None
            if val is None or val == "":
                continue

            minute = hcol_to_min(hname)

            try:
                req = Decimal(str(val))
            except Exception:
                raise HTTPException(status_code=400, detail=f"Valor inválido en {hname} ({dia}): {val}")

            existing = (
                db.query(Requirement)
                .filter(
                    Requirement.campaign_id == campaign_id,
                    Requirement.period == period,
                    Requirement.weekday == weekday,
                    Requirement.minute == minute,
                )
                .one_or_none()
            )

            if existing is None:
                db.add(
                    Requirement(
                        campaign_id=campaign_id,
                        period=period,
                        weekday=weekday,
                        minute=minute,
                        required=req,
                    )
                )
                inserted += 1
            else:
                if existing.required != req:
                    existing.required = req
                    updated += 1

            count_slots += 1

        weekday_slots[weekday] = weekday_slots.get(weekday, 0) + count_slots

    db.commit()

    return {
        "sheet": ws.title,
        "campaign_id": campaign_id,
        "period": period,
        "inserted": inserted,
        "updated": updated,
        "weekday_slots": weekday_slots,
    }


@app.get("/requirements")
def list_requirements(
    campaign_id: int,
    period: int,
    weekday: int | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(Requirement).filter(
        Requirement.campaign_id == campaign_id,
        Requirement.period == period,
    )
    if weekday is not None:
        q = q.filter(Requirement.weekday == weekday)

    rows = q.order_by(Requirement.weekday.asc(), Requirement.minute.asc()).all()
    return [
        {
            "campaign_id": r.campaign_id,
            "period": r.period,
            "weekday": r.weekday,
            "minute": r.minute,
            "required": str(r.required),
        }
        for r in rows
    ]
